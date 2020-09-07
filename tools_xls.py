

def find_row_col(sheet_data, target, startswith=False, index_0_based=True, col_chars_format=False):
    '''
        find target's row and col in sheet. 
        sheet_data is a 2d array, or iterable by rows, cols.
        startswith: find cell which is str and startswith(target)
        col_chars_format: return xls col name lick A,B...Z,AA,AB
        if not found, return None, None
    '''
    def find():
        for i, row in enumerate(sheet_data):
            for j, value in enumerate(row):
                if value==target:
                    return (i, j)
                if startswith and isinstance(value, str) and value.startswith(target):
                    return (i, j)
        return None, None

    row, col = find()
    if row is not None: # found
        col_chars = get_xls_column_letter(col+1)
        if not index_0_based:
            row += 1
            col += 1
        if col_chars_format:
            col = col_chars
    return row, col
        



def get_xlrd_cell_date(cell):
    import xlrd, datetime
    from tools_common import get_date_by_str
    if cell.ctype == 3: 
        ms_date_number = cell.value
        year, month, day, _hour, _minute, _second = xlrd.xldate_as_tuple(ms_date_number, xlrd.Book.datemode)
        res_date = datetime.date(year, month, day) 
        return res_date
    else:
        date_str = str(cell.value)
        return get_date_by_str(date_str)


def xls_column_index_from_string(chars):
    '''
        convert A->1, Z->26, AA->27, AB->28, AZZ->1378. indexx starts from 1
        the same as: openpyxl.utils.column_index_from_string
    '''
    ret = 0
    for char in chars:
        ret *= 26
        value = ord(char) - ord('A') + 1
        ret += value
    return ret


def get_xls_column_letter(indexx:int):
    '''
        convert 1->A, 26->Z, 17->AA, 28->AB, 1378->AZZ. indexx starts from 1
        the same as: from openpyxl.utils.get_column_letter
    '''
    ret = ''
    ord_zero = ord('A') - 1
    radix = ord('Z') - ord_zero
    while indexx > radix:
        rest = indexx % radix
        if (rest==0):
            ret += 'Z'
            indexx = indexx // radix - 1
        else:
            ret += chr(indexx % radix + ord_zero)
            indexx = indexx // radix
    if indexx > 0:
        ret += chr(indexx + ord_zero)
    return ret[::-1]



def openpyxl_usage():  
    # pylint: disable=unused-variable
    import openpyxl
    # creat or open WorkBook
    wb = openpyxl.Workbook() #create new
    wb = openpyxl.load_workbook('test.xls') #open current
    
    # create or open worksheet
    ws1 = wb.create_sheet("Mysheet1")  #insert to end
    ws2 = wb.create_sheet("Mysheet2", 0) #insert to head
    ws1 = wb["Mysheet1"]
    ws = wb.worksheets[0] # the first worksheet

    cell:openpyxl.cell.Cell = ws['A4']
    cell.value = 4
    ws['A4'] = 4 

    for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
        for cell in col:
            print(cell)
    cell_range = ws['A1':'C2']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]

    # ws.rows always starts form 1st row. 
    # ws.min_row is the minial row that contains data, but ws.rows doesn't start from ws.min_row
    print(ws.min_row, ws.max_row)
    print(ws.min_column, ws.max_column)
    print(openpyxl.utils.column_index_from_string('B'))
    print(openpyxl.utils.get_column_letter(2))

    wb.save('test.xls')


import xlwings
def find_xlwings_cell(rng:xlwings.Range, target, startswith=False):
    'cell is a 1x1 Range'
    for row in rng.rows:
        for cell in row:
            if (cell.value == target) or (startswith and cell.value.startswith(target)):
                return cell


def get_xlwings_col(cell:xlwings.Range):
    'cell is a 1x1 Range'
    sheet = cell.sheet
    for col in sheet.used_range.columns:
        if col.column == cell.column:
            return col


def get_xlwings_total_range(sheet:xlwings.Sheet) -> xlwings.Range:
    'return range from A1 to all used range'
    return sheet[:sheet.used_range.last_cell.row, :sheet.used_range.last_cell.column]


def xlwings_usage():
    import xlwings as xw
    wb = xw.Book()  # this will create a new workbook
    # wb = xw.Book('FileName.xlsx')  # connect to an existing file in the current working directory
    sheet:xlwings.Sheet = wb.sheets[0]  # sheet = wb.sheets['Sheet1']
    max_row = sheet.used_range.last_cell.row
    max_col = sheet.used_range.last_cell.column

    fill_ref_range = sheet[0, 3:max_col]
    fill_full_range = sheet[0:max_row, 0:max_col]
    # sheet.range('B103:AQ103').api.AutoFill(sheet.range('B103:AQ104').api,0)
    fill_ref_range.api.AutoFill(fill_full_range.api, 0)

    # an useful way to deal data:
    sheet = wb.sheets[0]
    product_name_col_index = find_xlwings_cell(sheet.used_range, 'product_name').colunm - 1
    equity_col_index = find_xlwings_cell(sheet.used_range, 'equity').colunm - 1
    equity_dict = {}
    for row in sheet.used_range.rows:
        product_name = sheet[row.row-1, product_name_col_index].value
        equity = sheet[row.row-1, equity_col_index].value
        if isinstance(product_name, str) and (type(equity) in [float, int]):
            equity_dict[product_name] = equity
    print(equity_dict)



